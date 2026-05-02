# -*- coding: utf-8 -*-
"""MOD-004 统一查询 — PIL全自动 + CMA/MSC手动输入"""
import asyncio, sys, os, json, shutil
from datetime import datetime
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')
from patchright.async_api import async_playwright

EVIDENCE = Path(r"D:\hermes\MODS\MOD004\data\evidence")
BACKUP = Path(r"D:\hermes\MODS\MOD004\data\backups")
EXCEL = Path(r"D:\hermes\2026-02-04海运动态.xlsx")
EVIDENCE.mkdir(parents=True, exist_ok=True)
BACKUP.mkdir(parents=True, exist_ok=True)

# ============================================================
# PIL 官网查询 (已验证可用)
# ============================================================
async def query_pil_batch(batch_list):
    """批量查询 PIL B/L"""
    results = {}
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page(viewport={"width":1920,"height":1080})
        
        # 先打开 PIL 首页
        await page.goto("https://www.pilship.com", timeout=45000, wait_until="domcontentloaded")
        await asyncio.sleep(5)
        
        # Cookie
        try:
            btn = await page.wait_for_selector('button:has-text("Deny")', timeout=3000)
            await btn.click()
            await asyncio.sleep(1)
        except: pass
        
        for item in batch_list:
            bl = item['bl_no']
            batch = item['batch_no']
            print(f"\nPIL: {batch} | B/L: {bl}")
            
            try:
                # 切换为 B/L Number
                await page.evaluate("""
                    document.querySelector('#customDropdownHome .custom-dropdown-selected')?.click();
                """)
                await asyncio.sleep(0.5)
                await page.evaluate("""
                    let opts = document.querySelectorAll('#customDropdownHome .custom-dropdown-options li');
                    for (let o of opts) {
                        if (o.getAttribute('data-value') === 'TrackTraceBL') { o.click(); break; }
                    }
                """)
                await asyncio.sleep(0.3)
                
                # 填入并搜索
                await page.fill('#refNoHome', '')
                await page.fill('#refNoHome', bl)
                await page.click('#homeCttSubmit')
                await asyncio.sleep(6)
                
                result_text = await page.inner_text("body")
                
                # 解析集装箱信息
                containers = _parse_pil_result(result_text)
                
                results[batch] = {
                    'bl': bl,
                    'containers': containers,
                    'status': 'ok' if containers else 'no_result',
                    'raw': result_text[:2000],
                }
                
                if containers:
                    latest = containers[0]
                    print(f"  {len(containers)}柜 | {latest['status']} @ {latest['location']} | {latest['datetime']}")
                else:
                    print(f"  无集装箱数据")
                
            except Exception as e:
                print(f"  错误: {e}")
                results[batch] = {'bl': bl, 'status': 'error', 'error': str(e)}
        
        await browser.close()
    return results


def _parse_pil_result(text):
    """解析 PIL 查询结果中的集装箱列表"""
    import re
    containers = []
    lines = text.split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if re.match(r'^(PCIU|TCLU|BMOU|TCNU|TRLU)\d+$', line):
            container_no = line
            if i+2 < len(lines) and lines[i+1].strip() == 'Trace':
                parts = lines[i+2].split()
                if len(parts) >= 5:
                    containers.append({
                        'no': container_no,
                        'type': parts[0] if parts else '',
                        'datetime': f"{parts[2]} {parts[3]}" if len(parts) > 3 else '',
                        'status': ' '.join(parts[4:-1]) if len(parts) > 5 else '',
                        'location': parts[-1] if parts else '',
                    })
                i += 2
        i += 1
    return containers


# ============================================================
# Excel 写回
# ============================================================
def update_excel(updates):
    """批量更新 Excel"""
    import openpyxl
    
    # 备份
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = BACKUP / f"海运动态_{ts}.xlsx"
    shutil.copy2(EXCEL, backup_path)
    print(f"备份: {backup_path}")
    
    wb = openpyxl.load_workbook(EXCEL)
    updated = 0
    
    for ws in wb.worksheets:
        for row in range(2, ws.max_row + 1):
            batch_cell = str(ws.cell(row=row, column=2).value or '')
            for batch_no, data in updates.items():
                if batch_no in batch_cell:
                    # 判断布局
                    hdr_e = str(ws.cell(row=1, column=5).value or '')
                    is_layout_b = any(kw in hdr_e.upper() for kw in ['PORT', '目的港', '起运港'])
                    
                    if is_layout_b:
                        eta_c, loc_c, next_c, st_c = 14, 15, 16, 17
                    else:
                        eta_c, loc_c, next_c, st_c = 13, 14, 15, 16
                    
                    if data.get('containers'):
                        c = data['containers'][0]
                        if 'Returned' in c['status']:
                            ws.cell(row=row, column=eta_c, value=f"ATA {c['datetime'][:9]}")
                            ws.cell(row=row, column=loc_c, value=f"空箱已归还({c['location'][:10]})")
                            ws.cell(row=row, column=next_c, value="全部柜已还空")
                        else:
                            ws.cell(row=row, column=loc_c, value=f"{c['status']} @ {c['location']}")
                            ws.cell(row=row, column=next_c, value=c['datetime'][:20])
                    
                    ws.cell(row=row, column=st_c, value=f"PIL查询完成")
                    updated += 1
                    print(f"  更新: {ws.title} Row {row} -> {batch_no}")
                    break
    
    wb.save(EXCEL)
    print(f"保存: {EXCEL} | 更新 {updated} 行")


# ============================================================
# CLI
# ============================================================
async def main():
    if len(sys.argv) < 2:
        print("用法:")
        print("  python query_final.py pil          # 批量查所有PIL")
        print("  python query_final.py pil BEW500   # 查指定B/L")
        print("  python query_final.py cma STATUS   # 手动输入CMA结果")
        sys.exit(1)
    
    cmd = sys.argv[1]
    
    # 从 Excel 读取待查列表
    from mod004.reader import read_all_sheets, filter_query_candidates
    sheets = read_all_sheets()
    
    if cmd == 'pil':
        candidates = [c for c in filter_query_candidates(sheets) if c.get('_carrier_norm') == 'PIL']
        
        if len(sys.argv) > 2:
            bl_filter = sys.argv[2].upper()
            candidates = [c for c in candidates if bl_filter in str(c.get('bl_no','')).upper()]
        
        print(f"PIL 待查询: {len(candidates)} 条")
        if len(candidates) > 10:
            print(f"限制批量 10 条 (安全起见)")
            candidates = candidates[:10]
        
        results = await query_pil_batch(candidates)
        
        # 写回 Excel
        update_excel({
            c['batch_no']: results.get(c['batch_no'], {'status': 'skipped'})
            for c in candidates
        })
    
    elif cmd == 'cma' or cmd == 'msc':
        print(f"\n{cmd.upper()} 需手动查询。请在浏览器打开跟踪页面输入提单号，复制结果后运行:")
        print(f"  python query_final.py {cmd} --manual")

asyncio.run(main())
