# MOD-004 M1: Excel 读取模块 (v2 — 支持多 Sheet 布局自动检测)
# 职责: 读取海运主表，按 Sheet 自动检测列布局，识别待查询 Booking 记录

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from .config import EXCEL_PATH, SKIP_QUERY_STATUSES, PROTECTED_COLUMNS, SUPPORTED_CARRIERS
from .layout import detect_layout, get_column_map_for_sheet, LAYOUT_TEMPLATES

log = logging.getLogger(__name__)


def _get_headers(ws) -> Dict[int, str]:
    """读取 Sheet 的表头行 (第1行)"""
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value is not None:
            headers[col_idx] = str(cell.value).strip()
    return headers


def read_all_sheets(filepath: Optional[Path] = None) -> Dict[str, List[Dict[str, Any]]]:
    """读取 Excel 所有 Sheet，按 Sheet 自动检测列布局"""
    path = filepath or EXCEL_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 跳过空 Sheet
        if ws.max_row <= 1:
            result[sheet_name] = []
            continue

        # 自动检测列布局
        headers = _get_headers(ws)
        col_map = get_column_map_for_sheet(headers)

        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            has_data = False

            for field, info in col_map.items():
                cell = ws.cell(row=row_idx, column=info["col"])
                value = cell.value
                if value is not None:
                    has_data = True
                    if info["type"] == str:
                        value = str(value).strip()
                    elif info["type"] == float and value != "":
                        try:
                            value = float(value)
                        except (ValueError, TypeError):
                            pass
                row_data[field] = value

            if has_data:
                row_data["_sheet"] = sheet_name
                row_data["_row"] = row_idx
                row_data["_layout"] = col_map  # 保留布局信息用于写回
                rows.append(row_data)

        result[sheet_name] = rows

    wb.close()
    log.info(f"读取完成: {len(result)} 个 Sheet, 总计 {sum(len(r) for r in result.values())} 行")
    return result


def filter_query_candidates(
    sheets: Dict[str, List[Dict[str, Any]]],
) -> List[Dict[str, Any]]:
    """筛选需要查询的记录"""
    candidates = []

    for sheet_name, rows in sheets.items():
        for row in rows:
            bl_no = row.get("bl_no")
            batch_no = row.get("batch_no")
            if not bl_no and not batch_no:
                continue

            carrier = row.get("carrier") or ""
            carrier_norm = carrier.upper().strip()
            if not carrier_norm or carrier_norm not in SUPPORTED_CARRIERS:
                continue

            status = str(row.get("query_status", "")).strip().upper()
            if any(s.upper() in status for s in SKIP_QUERY_STATUSES):
                continue

            row["_carrier_norm"] = carrier_norm
            candidates.append(row)

    log.info(f"待查询: {len(candidates)} 条记录")
    return candidates


def get_field(row: Dict[str, Any], field: str, default=None) -> Any:
    return row.get(field, default)


def is_field_empty(row: Dict[str, Any], field: str) -> bool:
    val = row.get(field)
    return val is None or (isinstance(val, str) and val.strip() == "")


def print_summary(sheets: Dict[str, List[Dict[str, Any]]]):
    """打印 Excel 读取摘要"""
    total = sum(len(r) for r in sheets.values())
    print(f"\n{'='*60}")
    print(f"  Excel 读取摘要")
    print(f"{'='*60}")
    for name, rows in sheets.items():
        carriers = set(str(r.get("carrier", "")).strip() for r in rows if r.get("carrier"))
        print(f"  Sheet: {name}  ({len(rows)} 行)")
        print(f"    船公司: {', '.join(sorted(carriers)) if carriers else '(无)'}")
        # 显示检测到的布局
        if rows:
            layout_info = rows[0].get("_layout", {})
            sample_cols = {k: v["col"] for k, v in list(layout_info.items())[:6]}
            print(f"    布局: 列映射 {sample_cols}")
    print(f"  总计: {total} 行")
    print(f"{'='*60}\n")
