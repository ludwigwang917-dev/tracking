# MOD-004 M6: Excel 写回模块 (v2 — 支持多 Sheet 布局)
# 职责: 将查询结果写回 Excel 主表，自动备份，保护字段不覆盖

import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional

import openpyxl

from .config import EXCEL_PATH, BACKUP_DIR, PROTECTED_COLUMNS

log = logging.getLogger(__name__)


def backup_excel(filepath: Optional[Path] = None) -> Path:
    """备份 Excel 文件"""
    src = filepath or EXCEL_PATH
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = BACKUP_DIR / f"{src.stem}_backup_{timestamp}{src.suffix}"
    shutil.copy2(src, dst)
    log.info(f"备份完成: {dst}")
    return dst


def write_updates(
    updates: List[Dict[str, Any]],
    filepath: Optional[Path] = None,
    dry_run: bool = False,
) -> Dict[str, int]:
    """批量写回 Excel 主表 (支持每行不同布局)

    Args:
        updates: [{_sheet, _row, _layout, field: new_value, ...}, ...]
    """
    path = filepath or EXCEL_PATH
    stats = {"written": 0, "skipped": 0, "protected": 0, "errors": 0}

    if not updates:
        return stats

    if not dry_run:
        backup_excel(path)

    wb = openpyxl.load_workbook(path)
    sheet_cache = {ws.title: ws for ws in wb.worksheets}

    for update in updates:
        sheet_name = update.get("_sheet", "")
        row_idx = update.get("_row", 0)
        layout = update.get("_layout", {})

        if not sheet_name or not row_idx or not layout:
            stats["skipped"] += 1
            continue

        ws = sheet_cache.get(sheet_name)
        if ws is None:
            stats["skipped"] += 1
            continue

        for field, new_value in update.items():
            if field.startswith("_"):
                continue
            if field not in layout:
                continue

            col_idx = layout[field]["col"]

            # 保护字段检查
            if field in PROTECTED_COLUMNS and not update.get("_force_write"):
                current_val = ws.cell(row=row_idx, column=col_idx).value
                if current_val is not None and str(current_val).strip():
                    stats["protected"] += 1
                    continue

            if not dry_run:
                ws.cell(row=row_idx, column=col_idx, value=new_value)
            stats["written"] += 1

    if not dry_run:
        wb.save(path)
        log.info(f"保存完成: {path}")

    wb.close()
    return stats
