# -*- coding: utf-8 -*-
"""
CMA / MSC 手动辅助查询
双击运行 → 弹出浏览器 → 在浏览器里查 → 回到终端按 Enter → 自动写回 Excel
"""
import subprocess, sys, os, json, shutil
from datetime import datetime
from pathlib import Path

EXCEL = Path(r"D:\hermes\2026-02-04海运动态.xlsx")
BACKUP = Path(r"D:\hermes\MODS\MOD004\data\backups")
BACKUP.mkdir(parents=True, exist_ok=True)

QUERIES = {
    "CMA": {
        "bl": "ABQ0117757",
        "batch": "GMS-C-851",
        "url": "https://www.cma-cgm.com/ebusiness/tracking",
    },
    "MSC": {
        "bl": "MEDUH7002122",
        "batch": "GMS-C-810",
        "url": "https://www.msc.com/track-a-shipment",
    },
}

RESULT_FILE = Path(r"D:\hermes\MODS\MOD004\data\manual_result.txt")


def open_browser_and_wait(carrier):
    """打开浏览器，等用户粘贴结果"""
    q = QUERIES[carrier]
    
    print(f"\n{'='*60}")
    print(f"  {carrier} 查询")
    print(f"{'='*60}")
    print(f"  Booking: {q['batch']}")
    print(f"  B/L:     {q['bl']}")
    print(f"")
    print(f"  >>> 浏览器应该已经打开: {q['url']}")
    print(f"  >>> 请输入提单号: {q['bl']}")
    print(f"  >>> 查看结果后，复制追踪信息粘贴到下面")
    print(f"  >>> 粘贴完按 Enter，输入空行结束")
    print(f"")
    
    # 尝试打开浏览器
    try:
        subprocess.Popen(['cmd', '/c', 'start', q['url']], shell=True)
    except:
        print("  (无法自动打开浏览器，请手动打开上面的 URL)")
    
    # 收集用户粘贴的追踪信息
    lines = []
    print("  请粘贴追踪结果 (空行结束):")
    while True:
        try:
            line = input()
            if not line.strip():
                break
            lines.append(line)
        except EOFError:
            break
    
    result_text = '\n'.join(lines)
    
    if not result_text.strip():
        print("\n  ⚠ 未输入任何内容，跳过")
        return None
    
    return parse_manual_input(carrier, result_text)


def parse_manual_input(carrier, text):
    """解析用户粘贴的追踪信息"""
    q = QUERIES[carrier]
    result = {
        "carrier": carrier,
        "batch_no": q["batch"],
        "bl": q["bl"],
        "raw_text": text,
        "containers": [],
    }
    
    import re
    # 找集装箱号 (4字母+7数字)
    found_containers = re.findall(r'\b([A-Z]{4}\d{7})\b', text)
    result["containers"] = list(set(found_containers))
    
    # 找状态关键词
    status_keywords = {
        "empty_returned": ["EMPTY", "RETURN", "归还", "还空", "空箱"],
        "delivered": ["DELIVERED", "DELIVERY", "已交付", "到港"],
        "in_transit": ["TRANSIT", "SAILING", "航行", "运输中", "UNDER WAY"],
        "loaded": ["LOADED", "LOADING", "装船", "已装"],
        "discharged": ["DISCHARGED", "DISCHARGE", "卸船", "卸货"],
    }
    
    text_upper = text.upper()
    for status, keywords in status_keywords.items():
        if any(kw in text_upper or kw in text for kw in keywords):
            result["status"] = status
            break
    
    # 找日期
    dates = re.findall(r'(\d{4}[-./]\d{1,2}[-./]\d{1,2}|\d{1,2}[-./]\d{1,2}[-./]\d{4})', text)
    result["dates"] = dates[:5]
    
    # 找地点
    locations = re.findall(r'(CHONGQING|SHANGHAI|TIANJIN|QINGDAO|NINGBO|YANTIAN|SINGAPORE|PORT\s*\w+|重庆|上海|天津|青岛)', text, re.IGNORECASE)
    result["locations"] = list(set(locations))[:5]
    
    return result


def update_excel(results):
    """写回 Excel"""
    import openpyxl
    
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = BACKUP / f"海运动态_{ts}.xlsx"
    shutil.copy2(EXCEL, backup_path)
    
    wb = openpyxl.load_workbook(EXCEL)
    
    for r in results:
        if not r:
            continue
        
        for ws in wb.worksheets:
            for row in range(2, ws.max_row + 1):
                batch_cell = str(ws.cell(row=row, column=2).value or '')
                if r['batch_no'] in batch_cell:
                    hdr_e = str(ws.cell(row=1, column=5).value or '')
                    is_layout_b = any(kw in hdr_e.upper() for kw in ['PORT', '目的港', '起运港'])
                    
                    eta_c, loc_c, next_c, st_c = (14, 15, 16, 17) if is_layout_b else (13, 14, 15, 16)
                    
                    summary = f"{r.get('status','')} | 柜数:{len(r.get('containers',[]))}"
                    if r.get('containers'):
                        summary += f" | {r['containers'][0]}"
                    
                    ws.cell(row=row, column=loc_c, value=summary[:50])
                    ws.cell(row=row, column=st_c, value=f"{r['carrier']}手动查询完成")
                    
                    print(f"  更新: {ws.title} Row {row} -> {r['batch_no']}")
                    break
    
    wb.save(EXCEL)
    print(f"保存: {EXCEL}")


def main():
    carrier = sys.argv[1].upper() if len(sys.argv) > 1 else None
    
    if carrier and carrier in QUERIES:
        result = open_browser_and_wait(carrier)
        if result:
            print(f"\n解析结果:")
            print(json.dumps(result, indent=2, ensure_ascii=False))
            with open(RESULT_FILE, 'a', encoding='utf-8') as f:
                f.write(json.dumps(result, ensure_ascii=False) + '\n')
            update_excel([result])
    else:
        print("CMA / MSC 手动查询工具")
        print("=" * 50)
        print("\n请选择:")
        print("  1 - 查询 CMA (ABQ0117757)")
        print("  2 - 查询 MSC (MEDUH7002122)")
        print("  3 - 两者都查")
        
        try:
            choice = input(">>> ").strip()
        except EOFError:
            choice = "3"
        
        results = []
        if choice in ('1', '3'):
            results.append(open_browser_and_wait("CMA"))
        if choice in ('2', '3'):
            results.append(open_browser_and_wait("MSC"))
        
        results = [r for r in results if r]
        if results:
            print(f"\n{'='*50}")
            print("解析结果:")
            for r in results:
                print(json.dumps(r, indent=2, ensure_ascii=False))
            update_excel(results)
            print("\n✓ 完成")


if __name__ == "__main__":
    main()
