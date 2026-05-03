# -*- coding: utf-8 -*-
"""CMA/MSC 自动查询 — 有头模式，你看着它自动操作"""
import asyncio, sys, os, json, shutil, re
from datetime import datetime
from pathlib import Path
from playwright.async_api import async_playwright

EXCEL = Path(r"D:\hermes\2026-02-04海运动态.xlsx")
BACKUP = Path(r"D:\hermes\MODS\MOD004\data\backups")
BACKUP.mkdir(parents=True, exist_ok=True)

QUERIES = {
    "CMA": {"bl": "ABQ0117757", "batch": "GMS-C-851",
            "url": "https://www.cma-cgm.com/ebusiness/tracking"},
    "MSC": {"bl": "MEDUH7002122", "batch": "GMS-C-810",
            "url": "https://www.msc.com/track-a-shipment"},
}


async def run_query(carrier):
    q = QUERIES[carrier]

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=['--no-sandbox', '--disable-blink-features=AutomationControlled']
        )
        page = await browser.new_page(viewport={"width": 1920, "height": 1080})

        # 隐藏自动化特征
        await page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => false});
            Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
            window.chrome = { runtime: {} };
        """)

        print(f"\n{'='*50}")
        print(f"  {carrier}: {q['bl']}")
        print(f"{'='*50}")

        # 打开跟踪页
        print(f"打开: {q['url']}")
        await page.goto(q['url'], timeout=60000, wait_until="domcontentloaded")
        await asyncio.sleep(8)

        title = await page.title()
        body = await page.inner_text("body")
        print(f"标题: {title}")
        print(f"页面: {len(body)} 字符 | {body[:300]}")

        # 处理 CMA 登录页
        if carrier == "CMA" and "cma-cgm" in title.lower() and len(body) < 100:
            print("\n⚠ CMA 是空白 SPA — 需要你手动操作一次:")
            print("  1. 在打开的浏览器中按 F5 刷新")
            print("  2. 如果跳出登录，登录 CMA 账号")
            print("  3. 确认跟踪页加载后，回到这里按 Enter")
            input("  >>> ")
            await page.reload()
            await asyncio.sleep(5)
            body = await page.inner_text("body")

        # 填入提单号
        print(f"\n填入提单号: {q['bl']}")
        filled = await page.evaluate(f"""
            (function() {{
                let inputs = document.querySelectorAll('input:not([type="hidden"]):not([type="submit"]):not([type="button"])');
                for (let inp of inputs) {{
                    if (inp.offsetParent !== null && inp.offsetWidth > 50) {{
                        inp.focus();
                        inp.value = '';
                        inp.value = '{q['bl']}';
                        inp.dispatchEvent(new Event('input', {{bubbles: true}}));
                        inp.dispatchEvent(new Event('change', {{bubbles: true}}));
                        return {{ok: true, id: inp.id, name: inp.name}};
                    }}
                }}
                return {{ok: false}};
            }})()
        """)
        print(f"  填入: {filled}")

        # 提交
        print("提交查询...")
        await page.keyboard.press("Enter")
        await asyncio.sleep(15)

        try:
            await page.wait_for_load_state("networkidle", timeout=20000)
        except:
            pass
        await asyncio.sleep(3)

        # 获取结果
        result = await page.inner_text("body")
        print(f"\n结果 ({len(result)} 字符):")
        print(result[:3000])

        # 解析并写回
        update_excel(carrier, q["batch"], result)

        print("\n按 Enter 关闭浏览器...")
        try:
            input()
        except:
            pass
        await browser.close()


def update_excel(carrier, batch_no, text):
    """解析结果并写回 Excel"""
    import openpyxl

    containers = list(set(re.findall(r'\b([A-Z]{4}\d{7})\b', text)))
    dates = re.findall(r'(\d{4}[-.]\d{2}[-.]\d{2}|\d{2}[-.]\d{2}[-.]\d{4})', text)

    status = ""
    s = text.lower()
    if "empty return" in s or "returned" in s:
        status = "空箱已归还"
    elif "discharged" in s or "卸" in text:
        status = "已卸船"
    elif "loaded" in s or "装" in text:
        status = "已装船"
    elif "gate out" in s or "delivery" in s:
        status = "已提柜"

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    shutil.copy2(EXCEL, BACKUP / f"海运动态_{ts}.xlsx")

    wb = openpyxl.load_workbook(EXCEL)
    for ws in wb.worksheets:
        for row in range(2, ws.max_row + 1):
            if batch_no in str(ws.cell(row=row, column=2).value or ''):
                hdr_e = str(ws.cell(row=1, column=5).value or '')
                is_b = any(kw in hdr_e.upper() for kw in ['PORT', '目的港'])
                eta_c, loc_c, next_c, st_c = (14, 15, 16, 17) if is_b else (13, 14, 15, 16)

                summary = f"{status} | {len(containers)}柜"
                if containers:
                    summary += f" | {containers[0]}"

                ws.cell(row=row, column=loc_c, value=summary[:50])
                ws.cell(row=row, column=st_c, value=f"{carrier}查询完成")
                print(f"  ✓ Excel: {ws.title} Row {row} -> {summary}")

    wb.save(EXCEL)
    print(f"  ✓ 保存: {EXCEL}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python query_chrome.py CMA")
        print("      python query_chrome.py MSC")
        sys.exit(1)
    asyncio.run(run_query(sys.argv[1].upper()))
