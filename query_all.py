# -*- coding: utf-8 -*-
"""MOD-004 一键批量查询 — PIL(无头) + CMA/MSC(有头Chrome)"""
import asyncio, sys, os, re, json, shutil
from datetime import datetime
from pathlib import Path
from playwright.async_api import async_playwright

EXCEL = Path(r"D:\hermes\2026-02-04海运动态.xlsx")
BACKUP = Path(r"D:\hermes\MODS\MOD004\data\backups")
BACKUP.mkdir(parents=True, exist_ok=True)

async def query_pil(page, bl_no, batch_no):
    """PIL 无头查询"""
    print(f"  PIL: {batch_no} | {bl_no}")
    await page.goto("https://www.pilship.com", timeout=30000, wait_until="domcontentloaded")
    await asyncio.sleep(3)
    
    # 切换 B/L
    await page.evaluate("""
        let sel = document.querySelector('#customDropdownHome .custom-dropdown-selected');
        if (sel) sel.click();
    """)
    await asyncio.sleep(0.5)
    await page.evaluate("""
        let opts = document.querySelectorAll('#customDropdownHome .custom-dropdown-options li');
        for (let o of opts) {
            if (o.getAttribute('data-value') === 'TrackTraceBL') { o.click(); break; }
        }
    """)
    await asyncio.sleep(0.3)
    
    await page.fill('#refNoHome', '')
    await page.fill('#refNoHome', bl_no)
    await page.click('#homeCttSubmit')
    await asyncio.sleep(6)
    
    result = await page.inner_text("body")
    containers = list(set(re.findall(r'\b(PCIU|TCLU|BMOU|TCNU|TRLU)\d+\b', result)))
    
    status = "已还空" if "EMPTY" in result.upper() else "在途"
    return {"carrier": "PIL", "batch": batch_no, "containers": containers, "status": status}


async def query_cma(page, bl_no, batch_no):
    """CMA 有头查询"""
    print(f"  CMA: {batch_no} | {bl_no}")
    await page.goto("https://www.cma-cgm.com/ebusiness/tracking", timeout=60000, wait_until="domcontentloaded")
    await asyncio.sleep(8)
    
    await page.evaluate(f"""
        let inp = document.querySelector('#Reference');
        if (inp) {{ inp.value = '{bl_no}'; inp.dispatchEvent(new Event('input', {{bubbles: true}})); }}
    """)
    await page.keyboard.press("Enter")
    await asyncio.sleep(12)
    
    result = await page.inner_text("body")
    containers = list(set(re.findall(r'\b([A-Z]{4}\d{7})\b', result)))
    status = "已还空" if "EMPTY IN DEPOT" in result else "DELIVERED" if "DELIVERED" in result else "在途"
    return {"carrier": "CMA", "batch": batch_no, "containers": containers, "status": status}


async def query_msc(page, bl_no, batch_no):
    """MSC 有头查询"""
    print(f"  MSC: {batch_no} | {bl_no}")
    await page.goto("https://www.msc.com/track-a-shipment", timeout=60000, wait_until="domcontentloaded")
    await asyncio.sleep(6)
    
    await page.evaluate(f"""
        let inputs = document.querySelectorAll('input');
        for (let inp of inputs) {{
            if (inp.offsetParent !== null) {{ inp.value = '{bl_no}'; break; }}
        }}
    """)
    await page.keyboard.press("Enter")
    await asyncio.sleep(15)
    
    result = await page.inner_text("body")
    containers = list(set(re.findall(r'\b([A-Z]{4}\d{7})\b', result)))
    return {"carrier": "MSC", "batch": batch_no, "containers": containers, "status": "查询完成"}


def update_excel(results):
    import openpyxl
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    shutil.copy2(EXCEL, BACKUP / f"海运动态_{ts}.xlsx")
    wb = openpyxl.load_workbook(EXCEL)
    updated = 0
    
    for r in results:
        for ws in wb.worksheets:
            for row in range(2, ws.max_row + 1):
                if r['batch'] in str(ws.cell(row=row, column=2).value or ''):
                    hdr_e = str(ws.cell(row=1, column=5).value or '')
                    is_b = any(kw in hdr_e.upper() for kw in ['PORT', '目的港'])
                    loc_c, st_c = (15, 17) if is_b else (14, 16)
                    
                    summary = f"{r['status']} | {len(r['containers'])}柜"
                    ws.cell(row=row, column=loc_c, value=summary)
                    ws.cell(row=row, column=st_c, value=f"{r['carrier']}查询完成")
                    updated += 1
    
    wb.save(EXCEL)
    print(f"\n✓ 更新 {updated} 行 | 备份: {backup.name}")


async def main():
    print("="*50)
    print("  MOD-004 一键批量查询")
    print("="*50)
    
    # 读取 Excel
    from mod004.reader import read_all_sheets, filter_query_candidates
    sheets = read_all_sheets()
    candidates = filter_query_candidates(sheets)
    
    # 取每个船司 1 条（可改）
    queries = {}
    for c in candidates:
        carrier = c.get('_carrier_norm', '')
        if carrier not in queries:
            queries[carrier] = {
                'batch_no': c['batch_no'],
                'bl_no': str(c.get('bl_no', '')),
            }
    
    print(f"待查: {queries}")
    
    results = []
    
    # PIL — 无头
    if 'PIL' in queries:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            
            for q in [queries['PIL']]:
                r = await query_pil(page, q['bl_no'], q['batch_no'])
                results.append(r)
            
            await browser.close()
    
    # CMA/MSC — 有头（同一个浏览器窗口顺序查）
    cmamsc = {k: v for k, v in queries.items() if k in ('CMA', 'MSC')}
    if cmamsc:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=False)
            page = await browser.new_page(viewport={"width": 1920, "height": 1080})
            await page.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {get: () => false});
                Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
                window.chrome = { runtime: {} };
            """)
            
            for carrier, q in cmamsc.items():
                if carrier == 'CMA':
                    r = await query_cma(page, q['bl_no'], q['batch_no'])
                else:
                    r = await query_msc(page, q['bl_no'], q['batch_no'])
                results.append(r)
            
            await browser.close()
    
    # 写回 Excel
    update_excel(results)
    
    # 打印摘要
    print("\n" + "="*50)
    for r in results:
        print(f"  {r['carrier']}: {r['status']} | {len(r['containers'])}柜")
        if r['containers']:
            print(f"    箱号: {r['containers'][:3]}..." if len(r['containers'])>3 else f"    箱号: {r['containers']}")
    print("="*50)


if __name__ == "__main__":
    asyncio.run(main())
