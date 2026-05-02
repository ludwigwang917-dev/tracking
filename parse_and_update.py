# -*- coding: utf-8 -*-
"""解析 PIL 查询结果并写回 Excel"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8')

# 读取查询结果
with open(r"D:\hermes\MODS\MOD004\data\evidence\pil_result_BL.txt", "r", encoding="utf-8") as f:
    text = f.read()

# 解析集装箱信息
containers = []
pattern = r'([A-Z0-9]+)\s*\nTrace\s*\n\s*(20GP|40GP|40HQ|40HC)\s*(FCL/FCL|LCL)\s*(\d{2}-[A-Za-z]{3}-\d{4}\s*\d{2}:\d{2}:\d{2})\s*(.+?)\s*([A-Z\s]+)\s*$'
# 用更简单的模式
lines = [l.strip() for l in text.split('\n') if l.strip()]

# 跳过标题和导航
data_start = False
for line in lines:
    if line.startswith('PCIU'):
        data_start = True
    if data_start:
        print(line)

# 找关键信息
# 找 container 号 + trace
container_pattern = re.findall(
    r'(PCIU\d+|TCLU\d+|BMOU\d+|TCNU\d+|TRLU\d+)\s*\nTrace\s*\n\s*(\w+)\s+(\S+)\s+(\d{2}-\w{3}-\d{4}\s+\d{2}:\d{2}:\d{2})\s+(.+?)\s+(\S+(?:\s+\S+)?)\s*\n',
    text
)

# 手动解析
i = 0
containers = []
while i < len(lines):
    line = lines[i]
    # 容器号以已知前缀开头
    if re.match(r'^(PCIU|TCLU|BMOU|TCNU|TRLU|BMOU)\d+$', line):
        container_no = line
        if i+1 < len(lines) and lines[i+1] == 'Trace':
            if i+2 < len(lines):
                parts = lines[i+2].split()
                if len(parts) >= 5:
                    container_type = parts[0]
                    cargo_type = parts[1]
                    date_time = f"{parts[2]} {parts[3]}"
                    # 状态和地点在后续
                    status = ' '.join(parts[4:-1]) if len(parts) > 5 else ''
                    location = parts[-1] if parts else ''
                    
                    containers.append({
                        'no': container_no,
                        'type': container_type,
                        'cargo': cargo_type,
                        'datetime': date_time,
                        'status': status,
                        'location': location,
                    })
        i += 3
    i += 1

print(f"\n解析到 {len(containers)} 个集装箱:")
for c in containers:
    print(f"  {c['no']} | {c['type']} | {c['datetime']} | {c['status']} | {c['location']}")

# 汇总信息
if containers:
    latest = containers[0]
    status_text = f"{latest['status']} @ {latest['location']}"
    eta_text = f"ATA{latest['datetime'][:9]}"  # 实际到港日
    
    print(f"\n汇总:")
    print(f"  总箱数: {len(containers)} × {containers[0]['type']}")
    print(f"  最新状态: {status_text}")
    print(f"  到港日期: {eta_text}")
    
    # 写回 Excel
    import openpyxl
    from pathlib import Path
    import shutil
    from datetime import datetime
    
    excel_path = Path(r"D:\hermes\2026-02-04海运动态.xlsx")
    backup_dir = Path(r"D:\hermes\MODS\MOD004\data\backups")
    backup_dir.mkdir(parents=True, exist_ok=True)
    
    # 备份
    backup = backup_dir / f"海运动态_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy2(excel_path, backup)
    print(f"\n备份: {backup}")
    
    # 打开并更新
    wb = openpyxl.load_workbook(excel_path)
    
    # 找到 GMS-C-616-15 所在行
    for ws in wb.worksheets:
        for row in range(2, ws.max_row + 1):
            batch = str(ws.cell(row=row, column=2).value or '')
            if 'GMS-C-616-15' in batch:
                # 获取布局信息
                # 判断是 layout_a 还是 layout_b
                header_E = str(ws.cell(row=1, column=5).value or '')
                if '目的港' in header_E or 'PORT' in header_E.upper():
                    # layout_b (铬)
                    eta_col = 14
                    loc_col = 15
                    next_col = 16
                    status_col = 17
                else:
                    # layout_a (萤石)
                    eta_col = 13
                    loc_col = 14
                    next_col = 15
                    status_col = 16
                
                # 写入
                ws.cell(row=row, column=eta_col, value=f"ATA 2026.04.20")
                ws.cell(row=row, column=loc_col, value=f"空箱已归还(CQ)")
                ws.cell(row=row, column=next_col, value="全部柜已还空")
                ws.cell(row=row, column=status_col, value="PIL查询完成")
                
                print(f"\n✓ 已更新 Excel:")
                print(f"  Sheet: {ws.title}, Row: {row}")
                print(f"  ETA → ATA 2026.04.20")
                print(f"  现在位置 → 空箱已归还(CQ)")
                print(f"  下一步 → 全部柜已还空")
                print(f"  查询状态 → PIL查询完成")
                break
    
    wb.save(excel_path)
    print(f"\n保存: {excel_path}")
